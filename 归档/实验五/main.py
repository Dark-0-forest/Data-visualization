#!/usr/bin/env python
# coding:utf-8
"""
Name    : main.py
Author  : Fu Ming
Time    : 2020/10/15 20:09
Desc    :
"""
import openpyxl
from pyecharts.charts import HeatMap
import pyecharts.globals as globals
from pyecharts import options as opts
from pyecharts.charts import Pie

globals._WarningControl.ShowWarning = False  # 关闭pyecharts给出的警告


def getClassifyData():
    filename = "classifyday1.xlsx"
    ws = openpyxl.load_workbook(filename)['classifyday1']
    classifyData = []
    for i in range(2, ws.max_row + 1):
        row = str(i)
        classifyData.append([ws['A' + row].value, ws['B' + row].value])
    return classifyData


def genHeatBaseData():
    x, y = [], []
    for i in range(100):
        fmt = str("{:0>2d}".format(i))
        x.append(fmt)
        y.append('1' + fmt)
    return x, y


def classifyHeatMap(xAxisData: list, yAxisData: list, data: list):
    HeatMap(init_opts=opts.InitOpts(width="2000px", height="2000px")).add_xaxis(xaxis_data=xAxisData).add_yaxis(
        series_name="Punch Card",
        yaxis_data=yAxisData,
        value=data,
        label_opts=opts.LabelOpts(
            is_show=False, color="#fff", position="bottom", horizontal_align="50%"
        ),
    ).set_series_opts().set_global_opts(
        legend_opts=opts.LegendOpts(is_show=True),
        xaxis_opts=opts.AxisOpts(
            type_="category",
            splitarea_opts=opts.SplitAreaOpts(
                is_show=True, areastyle_opts=opts.AreaStyleOpts(opacity=1)
            ),
            axislabel_opts=opts.LabelOpts(
                interval=0
            )
        ),
        yaxis_opts=opts.AxisOpts(
            type_="category",
            splitarea_opts=opts.SplitAreaOpts(
                is_show=True, areastyle_opts=opts.AreaStyleOpts(opacity=1)
            ),
        ),
        visualmap_opts=opts.VisualMapOpts(
            min_=0, max_=10, is_calculable=True, orient="horizontal", pos_left="center"
        ),
    ).render("classifyHeatMap.html")


def serializeClassifyData(data: list):
    job = {'waiter': 2, 'vip': 4, 'participant': 6, 'meeting': 8, 'reporter': 10}
    sData = {}
    for person in data:
        if not sData.__contains__(str(person[0])[0:3]):
            sData[str(person[0])[0:3]] = {}
        sData[str(person[0])[0:3]][str(person[0])[3:5]] = job[person[1]]
    return sData


def genClassifyHeatMapData(data: dict):
    rdata = []  # data中的每个列表内容依次是横坐标 纵坐标 值
    for k, v in data.items():
        for sk, sv in v.items():
            rdata.append([int(sk), int(k) - 100, sv])
    return rdata


def drawClassifyHeatMap():
    classifyData = getClassifyData()
    serializedClassifyData = serializeClassifyData(classifyData)
    xAxisValue, yAxisValue = genHeatBaseData()
    classifyHeatMapData = genClassifyHeatMapData(serializedClassifyData)
    classifyHeatMap(xAxisValue, yAxisValue, classifyHeatMapData)


# 计算每种人去了每个房间的总和
def getRoomData():
    classifyData = getClassifyData()
    serializedClassifyData = serializeClassifyData(classifyData)
    filename = "time_allocate_day1.xlsx"
    ws = openpyxl.load_workbook(filename)['time_allocate_day1']
    roomData = [[0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0]]
    for i in range(2, ws.max_row + 1):
        row = str(i)
        for room in range(6):
            roomData[((serializedClassifyData[str(ws['A' + row].value)[0:3]][
                str(ws['A' + row].value)[3:5]]) // 2) - 1][room] += ws[chr(103 + room) + row].value
    # roomData=[[0, 0, 0, 0, 60831, 234033], [89477, 1235993, 30267, 0, 60746, 0], [78857, 0, 30304, 0, 3632303, 0],
    # [633270, 0, 410705, 0, 0, 0], [3233, 1550, 2401, 119402, 0, 0]]
    return roomData


# 计算每个房间都被多少人呆过的总和
def getRoomData2(roomData: list):
    roomData2 = [[], [], [], [], [], []]
    for i in roomData:
        k = 0
        for v in i:
            roomData2[k].append(v)
            k += 1
    # roomData2=[[0, 89477, 78857, 633270, 3233], [0, 1235993, 0, 0, 1550], [0, 30267, 30304, 410705, 2401],[0, 0, 0,
    # 0, 119402], [60831, 60746, 3632303, 0, 0], [234033, 0, 0, 0, 0]]
    return roomData2


def drawRoomData2Pie():
    roomData2 = getRoomData2(getRoomData())
    jobs = ['waiter', 'vip', 'participant', 'meeting', 'reporter']
    newData = [[], [], [], [], [], []]
    k = 0
    for room in roomData2:
        i = 0
        for job in room:
            newData[k].append([jobs[i], job])
            i += 1
        k += 1
    k = 0
    pieCenter = [["20%", "30%"], ["55%", "30%"], ["85%", "30%"], ["20%", "70%"], ["55%", "70%"], ["85%", "70%"]]
    pie = Pie()
    for data in newData:
        pie.add("Room" + str(k + 1), data, center=pieCenter[k], radius=[40, 70])
        k += 1
    pie.set_global_opts(
        title_opts=opts.TitleOpts(title="Room1-6各人员访问量"),
        legend_opts=opts.LegendOpts(is_show=False),
    ).set_series_opts(label_opts=opts.LabelOpts(formatter="{b}: {c}", is_show=True))
    pie.render("roomTime.html")


# 1:休息区 2:嘉宾休息区 3:休息区 4:记者区 5:黑客竞赛现场 6：工作人员休息区
drawClassifyHeatMap()
drawRoomData2Pie()
